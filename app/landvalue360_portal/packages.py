from __future__ import annotations

import hashlib
import io
import json
import zipfile
from copy import deepcopy
from datetime import date
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .calculations import calculate_project
from .models import Project, ProjectDocument, ProjectVersion, User, utcnow
from .services import snapshot_from_db

PORTAL_FORMAT = "LANDVALUE360_PORTAL_SUBMISSION"
PORTAL_SCHEMA_VERSION = "portal-submission-1.0.0"
INTERNAL_FORMAT = "LANDVALUE360_PROJECT_PACKAGE"
INTERNAL_FORMAT_VERSION = "2.1.1"


def canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")


def sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _zip(files: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for name in sorted(files):
            archive.writestr(name, files[name])
    return buffer.getvalue()


def portal_submission_payload(db: Session, project: Project, version: ProjectVersion, applicant: User) -> dict[str, Any]:
    snapshot = snapshot_from_db(db, version)
    checks = calculate_project(snapshot)
    documents = list(db.scalars(select(ProjectDocument).where(
        ProjectDocument.project_id == project.id,
        ProjectDocument.project_version_id == version.id,
        ProjectDocument.deleted_at.is_(None),
    )).all())
    return {
        "schema_version": PORTAL_SCHEMA_VERSION,
        "package_version": "1.0.0",
        "project_id": project.id,
        "project_version_id": version.id,
        "project_reference": project.reference,
        "project_version_number": version.version_number,
        "submitted_at": (version.submitted_at or utcnow()).isoformat(),
        "applicant": {"user_id": applicant.id, "full_name": applicant.full_name, "email": applicant.email},
        "organization": {"organization_id": project.organization_id},
        "project_identity": {
            "name": project.name,
            "description": project.description,
            "currency": (snapshot.get("identity") or {}).get("currency") or "USD",
        },
        "land_data": deepcopy(snapshot.get("land") or {}),
        "planning_data": deepcopy(snapshot.get("planning") or {}),
        "land_uses": deepcopy(snapshot.get("land_uses") or []),
        "products": deepcopy(snapshot.get("products") or []),
        "pricing": [{
            "product_code": row.get("code"), "unit_selling_price": row.get("unit_selling_price"),
            "currency": row.get("currency"), "price_source": row.get("price_source"),
            "evidence_confidence": row.get("evidence_confidence"),
        } for row in snapshot.get("products") or []],
        "costs": deepcopy(snapshot.get("costs") or []),
        "documents_manifest": [{
            "document_id": row.id, "category": row.category, "original_name": row.original_name,
            "mime_type": row.mime_type, "size_bytes": row.size_bytes, "sha256": row.sha256,
        } for row in documents],
        "declarations": {
            "professional_notice": "Inputs are supplied by the applicant and require independent technical, financial and legal review.",
            "advanced_results_included": False,
        },
        "calculation_checks": checks,
        "currency": (snapshot.get("identity") or {}).get("currency") or "USD",
        "units": {"area": "sqm", "percentage": "percent", "money": "project_currency"},
    }


def portal_package_schema() -> dict[str, Any]:
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": "https://landvalue360.example/schemas/portal-submission-1.0.0.json",
        "title": "LandValue360 Portal Submission Package",
        "type": "object",
        "required": ["schema_version", "project_id", "project_version_id", "project_reference", "land_data", "planning_data", "land_uses", "products", "costs", "calculation_checks"],
        "properties": {
            "schema_version": {"const": PORTAL_SCHEMA_VERSION},
            "project_id": {"type": "string", "format": "uuid"},
            "project_version_id": {"type": "string", "format": "uuid"},
            "land_uses": {"type": "array"},
            "products": {"type": "array"},
            "costs": {"type": "array"},
        },
        "additionalProperties": True,
    }


def export_portal_package(db: Session, project: Project, version: ProjectVersion, applicant: User) -> bytes:
    submission = portal_submission_payload(db, project, version, applicant)
    declarations = submission["declarations"]
    documents = submission["documents_manifest"]
    files = {
        "submission.json": canonical_json(submission),
        "schema.json": canonical_json(portal_package_schema()),
        "documents-manifest.json": canonical_json(documents),
        "declarations.json": canonical_json(declarations),
    }
    manifest = {
        "format": PORTAL_FORMAT,
        "format_version": "1.0.0",
        "schema_version": PORTAL_SCHEMA_VERSION,
        "exported_at_utc": utcnow().isoformat(),
        "project_reference": project.reference,
        "project_version": version.version_number,
        "files": {name: {"sha256": sha(data), "bytes": len(data)} for name, data in files.items()},
    }
    files["manifest.json"] = canonical_json(manifest)
    files["checksums.json"] = canonical_json({name: sha(data) for name, data in files.items()})
    return _zip(files)


def _curve(start_year: int, duration_months: int | None) -> list[dict[str, str]]:
    years = max(1, min(8, round((duration_months or 36) / 12)))
    weights = [1 / years] * years
    return [{"date": f"{start_year + i}-01-01", "weight": str(round(w, 10))} for i, w in enumerate(weights)]


def internal_input_snapshot(project: Project, version: ProjectVersion, *, source_snapshot: dict[str, Any] | None = None) -> dict[str, Any]:
    """Build the native Platform 2.1.1 detailed-input contract.

    The Client Portal captures product allocation, sellable efficiency and
    commercial pricing in one user-facing row.  The internal Platform keeps
    planning allocation in ``planning_products`` and pricing/timing in
    ``products``.  The package therefore performs that explicit split while
    retaining every original row in ``portal_source`` and
    ``source_input_snapshot`` for audit.
    """
    snapshot = deepcopy(source_snapshot if source_snapshot is not None else (version.input_snapshot or {}))
    identity = snapshot.get("identity") or {}
    land = snapshot.get("land") or {}
    planning = snapshot.get("planning") or {}
    planning_products: list[dict[str, Any]] = []
    products: list[dict[str, Any]] = []
    model_valuation_date = ((snapshot.get("financial_model") or {}).get("valuation_date") or date.today().isoformat())
    try:
        valuation_date = date.fromisoformat(str(model_valuation_date)).isoformat()
    except ValueError:
        valuation_date = date.today().isoformat()
    start_year = date.fromisoformat(valuation_date).year
    for index, row in enumerate(snapshot.get("products") or [], start=1):
        product_id = str(row.get("code") or f"PORTAL-PRODUCT-{index}").strip()
        name = row.get("name") or product_id
        allocation_share = str(float(row.get("allocation_percentage") or 0) / 100)
        efficiency = str(float(row.get("sellable_efficiency_percentage") or 0) / 100)
        planning_products.append({
            "product_id": product_id,
            "name": name,
            "area_method": "GFA_ALLOCATION",
            "gfa_allocation_share": allocation_share,
            "efficiency": efficiency,
            "is_sellable": True,
        })
        products.append({
            "product_id": product_id,
            "name": name,
            "quantity_basis": "SELLABLE_AREA_SQM",
            "quantity_unit": "sqm",
            "unit_price": str(row.get("unit_selling_price") or 0),
            "description": "Imported from LandValue360 Client Portal; analyst review required.",
            "market_growth_rate": "0",
            "payment_plan_id": "PORTAL_REVIEW",
            "construction_cost_per_sqm": "0",
            "sales_curve_type": "CUSTOM",
            "sales_start_month": 1,
            "sales_duration_months": planning.get("sales_duration_months") or 36,
            "construction_curve_type": "CUSTOM",
            "construction_start_month": 1,
            "construction_duration_months": planning.get("project_duration_months") or 36,
            "commercial_discount_rate": "0",
            "buyer_incentive_rate": "0",
            "refund_rate": "0",
            "buyer_incentive_net_sales_deduction_fraction": "1",
            "refund_net_sales_deduction_fraction": "1",
            "eligible_profit_share_revenue_fraction": "1",
            "construction_developer_responsibility_share": "1",
            "construction_government_responsibility_share": "0",
            "construction_developer_economic_share": "1",
            "construction_government_economic_share": "0",
            "sales_curve": _curve(start_year, planning.get("sales_duration_months")),
            "collection_rules": [{"lag_days": 0, "weight": "1"}],
            "portal_source": deepcopy(row),
        })
    costs = []
    for index, row in enumerate(snapshot.get("costs") or [], start=1):
        amount = row.get("amount")
        if amount in (None, ""):
            amount = str(float(row.get("quantity") or 0) * float(row.get("unit_cost") or 0))
        developer_share = float(row.get("developer_share_percentage") or 100) / 100
        costs.append({
            "cost_id": f"PORTAL-{index:03d}",
            "name": row.get("name") or f"Portal cost {index}",
            "category": row.get("category") or "CUSTOM",
            "quantity": str(row.get("quantity") or 1),
            "unit_cost": str(row.get("unit_cost") or amount or 0),
            "base_date": valuation_date,
            "escalation_rate": "0",
            "contingency_rate": "0",
            "developer_responsibility_share": str(developer_share),
            "government_responsibility_share": str(1 - developer_share),
            "developer_economic_share": str(developer_share),
            "government_economic_share": str(1 - developer_share),
            "eligible_net_sales_deduction_fraction": "1" if row.get("net_sales_deductible") else "0",
            "eligible_profit_share_cost_fraction": "1",
            "is_direct_cost": row.get("category") in {"CONSTRUCTION", "INFRASTRUCTURE", "PUBLIC_FACILITIES"},
            "expenditure_curve": _curve(start_year, planning.get("project_duration_months")),
            "calculation_method": "FIXED_AMOUNT",
            "fixed_amount": str(amount or 0),
            "calculation_note": "Imported portal estimate; timing, escalation and eligibility require analyst confirmation.",
            "portal_source": deepcopy(row),
        })
    return {
        "project_id": project.id,
        "project_name": project.name,
        "reporting_currency": identity.get("currency") or "USD",
        "valuation_date": valuation_date,
        "land_value_baseline": str(land.get("current_land_value") or 0),
        "reference_land_value_basis": "GROSS",
        "reference_land_value_area_sqm": str(land.get("gross_land_area_sqm") or 0),
        "reference_land_value_total": str(land.get("current_land_value") or 0),
        "planning": {
            "gross_land_area_sqm": str(land.get("gross_land_area_sqm") or 0),
            "excluded_land_area_sqm": str(land.get("excluded_land_area_sqm") or 0),
            "far_land_basis": "NET",
            "far": str(planning.get("far") or 0),
            "bcr_land_basis": "NET",
            "bcr": str(planning.get("bcr") or 0),
            "land_uses": [{"land_use_id": row.get("code"), "name": row.get("name"), "share": str(float(row.get("percentage") or 0) / 100)} for row in snapshot.get("land_uses") or []],
        },
        "planning_products": planning_products,
        "products": products,
        "costs": costs,
        "funding": {"opening_cash": "0", "committed_additional_equity": "0", "committed_equity": "0", "committed_equity_is_additional": True, "committed_financing": "0"},
        "finance_model": {"enabled": False, "annual_interest_rate": "0", "minimum_cash_balance": "0", "spend_policy": "SCHEDULE_DRIVEN", "future_cost_reserve_share": "0", "allow_negative_cash": True, "defer_contractual_payments": True, "maximum_extension_months": 120},
        "partnership": {"method": "GROSS_SALES", "share_rate": "0", "approved_selection": "MANUAL", "manual_share": "0", "net_deduction_treatment": "CUMULATIVE_CARRY_FORWARD", "upfront_payments": []},
        "portal_submission": {
            "schema_version": PORTAL_SCHEMA_VERSION,
            "source_platform_version": "financial-portal-2.5.0",
            "source_project_reference": project.reference,
            "source_project_version": version.version_number,
            "target_internal_contract": INTERNAL_FORMAT_VERSION,
            "requires_analyst_completion": True,
        },
    }


def export_internal_package(db: Session, project: Project, version: ProjectVersion) -> bytes:
    """Export the effective monthly financial input to Platform 2.1.1.

    The former client portal exported placeholder finance, cash and landowner
    assumptions.  The standalone financial portal instead materializes the
    policy-dependent financial defaults, normalizes the saved model and exports
    the exact Platform-native input contract consumed by its embedded engine.
    """
    # Local imports avoid a module cycle: the financial adapter itself reuses
    # ``internal_input_snapshot`` as the structural Platform contract builder.
    from .financial_engine import build_engine_project_snapshot, effective_project_input_snapshot
    from .financial_service import current_policy_version

    policy_version = current_policy_version(db)
    source_snapshot = effective_project_input_snapshot(version, policy_version.policy_snapshot)
    input_snapshot = build_engine_project_snapshot(
        project,
        version,
        source_snapshot["financial_model"],
        policy_version.policy_snapshot,
        source_snapshot=source_snapshot,
    )
    effective_input_hash = sha(canonical_json(input_snapshot))
    project_payload = {
        "name": project.name, "code": project.reference, "description": project.description,
        "project_kind": "SHARED", "status": "DRAFT", "source_project_id": project.id,
        "created_at": project.created_at.isoformat(), "updated_at": project.updated_at.isoformat(),
    }
    versions = [{
        "source_version_id": version.id,
        "version_number": version.version_number,
        "source_status": "PORTAL_SUBMITTED",
        "label": f"Portal submission v{version.version_number}",
        "notes": "Imported from the external portal. All advanced timing, financing, tax and contract assumptions require analyst review before calculation.",
        "input_snapshot": input_snapshot,
        "input_hash": effective_input_hash,
        "source_input_schema": PORTAL_SCHEMA_VERSION,
        "source_input_snapshot": deepcopy(version.input_snapshot),
        "source_input_hash": version.snapshot_hash,
        "supersedes_source_version_id": None,
    }]
    scenarios: list[dict[str, Any]] = []
    files = {
        "project.json": canonical_json(project_payload),
        "versions.json": canonical_json(versions),
        "scenarios.json": canonical_json(scenarios),
    }
    manifest = {
        "format": INTERNAL_FORMAT,
        "format_version": INTERNAL_FORMAT_VERSION,
        "source_platform_version": "financial-portal-2.5.0",
        "exported_at_utc": utcnow().isoformat(),
        "project": {"name": project.name, "code": project.reference, "project_kind": "SHARED"},
        "counts": {"versions": 1, "scenarios": 0, "reference_results": 0},
        "reference_results": [],
        "files": {name: {"sha256": sha(data), "bytes": len(data)} for name, data in files.items()},
        "compatibility": {
            "target_platform": "LandValue360 Platform 2.1.1 or newer",
            "native_detailed_contract": True,
            "planning_products_separated": True,
            "requires_recalculation": True,
            "monthly_financial_inputs_included": True,
            "policy_version_id": policy_version.id,
            "policy_snapshot_hash": policy_version.snapshot_hash,
            "effective_input_hash": effective_input_hash,
            "source_input_hash": version.snapshot_hash,
        },
        "import_rule": "Portal inputs import as a SHARED draft with the effective monthly finance, timing and landowner assumptions. Recalculate in LandValue360 Platform before using advanced reports.",
    }
    files["manifest.json"] = canonical_json(manifest)
    return _zip(files)


def export_excel(project: Project, version: ProjectVersion) -> bytes:
    snapshot = deepcopy(version.input_snapshot or {})
    result = calculate_project(snapshot)
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    rows = [
        ("Project", project.name), ("Reference", project.reference), ("Version", version.version_number),
        ("Net land area sqm", result["net_land_area_sqm"]), ("Total GFA sqm", result["total_gfa_sqm"]),
        ("Sellable area sqm", result["total_sellable_area_sqm"]), ("Gross sales", result["gross_sales_nominal"]),
        ("Total costs", result["total_costs"]), ("Developer costs", result["developer_costs"]),
        ("Landowner costs", result["landowner_costs"]), ("Deductible cost flags", result["net_sales_deductible_costs"]),
    ]
    for row in rows: ws.append(row)
    w = wb.create_sheet("Land Uses"); w.append(["Code", "Name", "%", "Area sqm"])
    for row in result["land_uses"]: w.append([row.get("code"), row.get("name"), row.get("percentage"), row.get("area_sqm")])
    w = wb.create_sheet("Products"); w.append(["Code", "Name", "% GFA", "GFA sqm", "Efficiency %", "Sellable sqm", "Unit price", "Gross sales"])
    for row in result["products"]: w.append([row.get("code"), row.get("name"), row.get("allocation_percentage"), row.get("product_gfa_sqm"), row.get("sellable_efficiency_percentage"), row.get("sellable_area_sqm"), row.get("unit_selling_price"), row.get("gross_sales")])
    w = wb.create_sheet("Costs"); w.append(["Name", "Category", "Amount", "Developer %", "Landowner %", "Deductible"])
    for row in result["costs"]: w.append([row.get("name"), row.get("category"), row.get("amount"), row.get("developer_share_percentage"), row.get("landowner_share_percentage"), row.get("net_sales_deductible")])
    buffer = io.BytesIO(); wb.save(buffer); return buffer.getvalue()
