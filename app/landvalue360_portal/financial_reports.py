from __future__ import annotations

from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from html import escape
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from .report_renderer import render_html_pdf

from .financial_engine import D, policy_controls, sha256_json
from .models import CalculationRun, EngineVersion, FinancialPolicyVersion, Project, ProjectVersion
from landvalue360_server.constraint_registry import constraint_metadata

DARK = "173F4A"
ACCENT = "1D5E5A"
LIGHT = "E8F1F2"
INPUT_BLUE = "0000FF"
STATIC_GRAY = "666666"
CAUTION = "FCE8C3"
ERROR = "FADBD8"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "C00000"


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _money(value: Any) -> str:
    number = _decimal(value)
    if number is None:
        return "-"
    return f"{number:,.0f}"


def _percent(value: Any) -> str:
    number = _decimal(value)
    if number is None:
        return "-"
    return f"{number * Decimal('100'):,.1f}%"


def _multiple(value: Any) -> str:
    number = _decimal(value)
    if number is None:
        return "-"
    return f"{number:,.2f}x"


def _value_by_type(value: Any, measure_type: str) -> str:
    return _money(value) if str(measure_type).upper() == "AMOUNT" else _percent(value)


def _constraint_value(constraint_id: Any, value: Any) -> str:
    number = _decimal(value)
    if number is None:
        return "-"
    code = str(constraint_id or "").upper()
    if "IRR" in code or "PROFIT_ON_COST" in code:
        return _percent(number)
    if "MULTIPLE" in code or "MOIC" in code:
        return _multiple(number)
    if any(token in code for token in ("NPV", "GAP", "DEBT", "SCOPE", "SHORTFALL", "ARREARS", "PROFIT_NONNEGATIVE")):
        return _money(number)
    if number == 0:
        return "0"
    absolute = abs(number)
    if absolute < Decimal("0.0001"):
        return f"{number:.2E}"
    if absolute < Decimal("10"):
        return f"{number:,.4f}"
    return f"{number:,.2f}"


def _set_widths(ws, widths: dict[int, float]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _section(ws, row: int, title: str, end_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.font = Font(color=WHITE, bold=True, size=11)
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], *, number_columns: set[int] | None = None, percent_columns: set[int] | None = None, multiple_columns: set[int] | None = None) -> int:
    number_columns = number_columns or set()
    percent_columns = percent_columns or set()
    multiple_columns = multiple_columns or set()
    thin = Side(style="thin", color="B8C7CA")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        cell.font = Font(bold=True, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    row_number = start_row + 1
    money_format = '#,##0;[Red](#,##0);-'
    percent_format = '0.0%;[Red](0.0%);-'
    multiple_format = '0.00x;[Red](0.00x);-'
    for values in rows:
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_number, col)
            parsed = _decimal(value) if col in number_columns | percent_columns | multiple_columns else None
            cell.value = float(parsed) if parsed is not None else value
            cell.alignment = Alignment(horizontal="right" if col in number_columns | percent_columns | multiple_columns else "left", vertical="center", wrap_text=True)
            if col in number_columns:
                cell.number_format = money_format
            elif col in percent_columns:
                cell.number_format = percent_format
            elif col in multiple_columns:
                cell.number_format = multiple_format
        row_number += 1
    return row_number


def build_financial_excel(
    *,
    project: Project,
    project_version: ProjectVersion,
    run: CalculationRun,
    policy_version: FinancialPolicyVersion,
    engine_version: EngineVersion,
    payload: dict[str, Any],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    summary = payload.get("summary") or {}
    truth = payload.get("financial_truth") or {}
    residual = payload.get("residual_valuation") or {}
    negotiation = payload.get("negotiation_results") or []
    annual = payload.get("annual_cashflow") or []
    monthly = payload.get("monthly_cashflow") or []
    selected_policy_controls = policy_controls(policy_version.policy_snapshot)
    policy_name_ar = str(selected_policy_controls.get("display_name_ar") or f"السياسة v{policy_version.version_number}")
    policy_name_en = str(selected_policy_controls.get("display_name_en") or f"Policy v{policy_version.version_number}")
    policy_description_ar = str(selected_policy_controls.get("description_ar") or "")
    policy_description_en = str(selected_policy_controls.get("description_en") or "")

    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    _set_widths(ws, {1: 34, 2: 22, 3: 34, 4: 22})
    ws.merge_cells("A1:D1")
    ws["A1"] = "LandValue360 - Standalone Financial Portal"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"{project.name} | {project.reference} | Project Version {project_version.version_number}"
    ws["A2"].font = Font(size=12, bold=True, color=DARK)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:D3")
    ws["A3"] = f"Run {run.id} | {policy_name_en} v{policy_version.version_number} | Engine {engine_version.engine_version}/{engine_version.adapter_version}"
    ws["A3"].font = Font(size=9, color=STATIC_GRAY)
    ws["A3"].alignment = Alignment(horizontal="center")
    row = 5
    row = _section(ws, row, "Core Feasibility Indicators", 4)
    metrics = [
        ("Gross Sales", summary.get("gross_sales"), "Net Sales", summary.get("net_sales")),
        ("Development Cost", summary.get("development_cost"), "Project Profit", summary.get("project_profit")),
        ("Developer Profit", summary.get("developer_profit"), "Peak Equity Requirement", summary.get("peak_equity")),
        ("Equity Contributions", summary.get("developer_equity_contributions"), "Equity Distributions", summary.get("developer_equity_distributions")),
        ("Peak Funding Gap", summary.get("peak_funding_gap"), "Peak Debt", summary.get("peak_debt")),
        ("Interest", summary.get("interest_total"), "Financing Fees", summary.get("financing_fees_total")),
        ("Ending Cash", summary.get("ending_cash"), "Terminal Debt", summary.get("terminal_debt")),
        ("Deferred Costs", summary.get("deferred_costs"), "Contractual Arrears", summary.get("contractual_arrears")),
    ]
    money_format = '#,##0;[Red](#,##0);-'
    for left_label, left_value, right_label, right_value in metrics:
        ws.cell(row, 1, left_label).font = Font(color=STATIC_GRAY)
        ws.cell(row, 2, float(D(left_value))).number_format = money_format
        ws.cell(row, 2).font = Font(bold=True, color=BLACK)
        ws.cell(row, 3, right_label).font = Font(color=STATIC_GRAY)
        ws.cell(row, 4, float(D(right_value))).number_format = money_format
        ws.cell(row, 4).font = Font(bold=True, color=BLACK)
        row += 1
    row += 1
    row = _section(ws, row, "Return Indicators", 4)
    returns = [
        ("Project Profit on Cost", summary.get("project_profit_on_cost"), "Project Profit on Revenue", summary.get("project_profit_on_revenue")),
        ("Developer Profit on Cost", summary.get("developer_profit_on_cost"), "Developer Profit on Revenue", summary.get("developer_profit_on_revenue")),
        ("Project IRR", summary.get("project_irr"), "Project NPV", summary.get("project_npv")),
        ("Developer Equity IRR", summary.get("developer_equity_irr"), "Developer Equity NPV", summary.get("developer_equity_npv")),
        ("Developer MOIC", summary.get("developer_equity_multiple"), "Policy Compliant", "YES" if payload.get("policy_compliant") else "NO"),
    ]
    for left_label, left_value, right_label, right_value in returns:
        ws.cell(row, 1, left_label).font = Font(color=STATIC_GRAY)
        left = ws.cell(row, 2)
        left.value = float(D(left_value)) if left_value not in (None, "") else None
        left.number_format = '0.00x' if "MOIC" in left_label else '0.0%'
        left.font = Font(bold=True)
        ws.cell(row, 3, right_label).font = Font(color=STATIC_GRAY)
        right = ws.cell(row, 4)
        if right_label.endswith("NPV"):
            right.value = float(D(right_value))
            right.number_format = money_format
        elif right_label == "Policy Compliant":
            right.value = right_value
            right.fill = PatternFill("solid", fgColor=LIGHT if right_value == "YES" else ERROR)
        else:
            right.value = float(D(right_value)) if right_value not in (None, "") else None
            right.number_format = '0.0%'
        right.font = Font(bold=True)
        row += 1
    row += 1
    row = _section(ws, row, "Project Duration", 4)
    duration_rows = [
        ("Original Duration (months)", summary.get("original_project_duration_months"), "Adjusted Duration (months)", summary.get("adjusted_project_duration_months")),
        ("Original Completion Date", summary.get("original_completion_date"), "Adjusted Completion Date", summary.get("adjusted_completion_date")),
    ]
    for left_label, left_value, right_label, right_value in duration_rows:
        ws.cell(row, 1, left_label).font = Font(color=STATIC_GRAY)
        ws.cell(row, 2, left_value).font = Font(bold=True)
        ws.cell(row, 3, right_label).font = Font(color=STATIC_GRAY)
        ws.cell(row, 4, right_value).font = Font(bold=True)
        row += 1
    row += 1
    row = _section(ws, row, "Residual Land Value", 4)
    residual_rows = [
        ("Gross Development Value", residual.get("gross_development_value")),
        ("Target Developer Profit on Cost", residual.get("target_developer_profit_on_cost")),
        ("Development Costs", residual.get("development_costs")),
        ("Finance Costs", residual.get("finance_costs")),
        ("Residual Land Value", residual.get("residual_land_value")),
        ("Land Capacity DCF", residual.get("land_capacity_dcf")),
    ]
    for label, value in residual_rows:
        ws.cell(row, 1, label)
        cell = ws.cell(row, 2, float(D(value)))
        cell.number_format = '0.0%' if label == "Target Developer Profit on Cost" else money_format
        if label in {"Residual Land Value", "Land Capacity DCF"}:
            cell.font = Font(bold=True, color=ACCENT)
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=4)
    ws.cell(row, 1, residual.get("label_en") or "Development Residual Indication - not an independent market valuation")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=CAUTION)
    ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")

    ws = wb.create_sheet("Negotiation Range")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    _set_widths(ws, {1: 22, 2: 20, 3: 18, 4: 18, 5: 22, 6: 22, 7: 22, 8: 20, 9: 24, 10: 20, 11: 12, 12: 22, 13: 22, 14: 18})
    headers = ["Mechanism", "Status", "Minimum Acceptable", "Balanced", "Policy-Adjusted Ceiling", "Technical Ceiling / Search Cap", "Residual Land Value", "Residual Equivalent", "Governing Constraint", "Offer Position", "Rank", "Landowner NPV @ Balanced", "Developer NPV @ Balanced", "Developer IRR @ Balanced"]
    rows = []
    for item in negotiation:
        balanced_case = item.get("balanced_case") or {}
        rows.append([
            item.get("method"), item.get("ceiling_kind") or item.get("status"), item.get("fair_floor"), item.get("balanced", item.get("recommended")), item.get("policy_adjusted_ceiling", item.get("risk_adjusted_ceiling")), item.get("technical_ceiling"),
            item.get("residual_land_value"), item.get("residual_equivalent_measure"), item.get("governing_constraint_id"), item.get("offer_position"), item.get("recommendation_rank"), balanced_case.get("government_npv", item.get("government_npv")), balanced_case.get("developer_npv", item.get("developer_npv")), balanced_case.get("developer_equity_irr", balanced_case.get("developer_irr", item.get("developer_irr"))),
        ])
    _write_table(ws, 1, headers, rows, number_columns={3, 4, 5, 6, 7, 8, 12, 13}, percent_columns={14})
    for row_idx, item in enumerate(negotiation, 2):
        if str(item.get("measure_type")).upper() == "RATE":
            for col in (3, 4, 5, 6, 8):
                ws.cell(row_idx, col).number_format = '0.00%'
        if str(item.get("status")) != "VALID_RANGE":
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=CAUTION)

    ws = wb.create_sheet("Annual Cash Flow")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = ["Year", "Gross Contracted Sales", "Gross Collections", "Net Collections", "Planned Cost", "Actual Cost", "Deferred Cost", "Equity", "Debt Draw", "Interest", "Finance Fees", "Debt Repayment", "Landowner Payments", "Developer Distributions", "Ending Cash", "Ending Debt", "Funding Gap", "Contractual Arrears"]
    rows = [[
        item.get("year"), item.get("gross_contracted_sales"), item.get("gross_collections"), item.get("net_collections"), item.get("planned_cost"), item.get("actual_cost"), item.get("deferred_cost"), item.get("equity_contribution"), item.get("financing_draw"), item.get("interest_paid"), item.get("financing_fees"), item.get("financing_repayment"), item.get("landowner_cash_receipt", item.get("government_payment")), item.get("developer_distribution"), item.get("ending_cash"), item.get("ending_debt"), item.get("unsupported_funding_gap"), item.get("contractual_arrears"),
    ] for item in annual]
    _write_table(ws, 1, headers, rows, number_columns=set(range(2, 19)))
    _set_widths(ws, {1: 12, **{index: 20 for index in range(2, 19)}})

    ws = wb.create_sheet("Monthly Cash Flow")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = ["Month", "Date", "Opening Cash", "Gross Contracted Sales", "Net Contracted Sales", "Gross Collections", "Net Collections", "Planned Cost", "Actual Cost", "Deferred Cost", "Equity Contribution", "Financing Draw", "Interest Paid", "Financing Fees", "Debt Repayment", "Landowner Payment", "Developer Distribution", "Ending Cash", "Ending Debt", "Funding Gap", "Contractual Arrears", "Cash Reconciliation Variance"]
    rows = [[
        item.get("month"), item.get("date"), item.get("opening_cash"), item.get("gross_contracted_sales"), item.get("net_contracted_sales"), item.get("gross_collections"), item.get("net_collections"), item.get("planned_cost"), item.get("actual_cost"), item.get("deferred_cost"), item.get("equity_contribution"), item.get("financing_draw"), item.get("interest_paid"), item.get("financing_fees"), item.get("financing_repayment"), item.get("landowner_cash_receipt", item.get("government_payment")), item.get("developer_distribution"), item.get("ending_cash"), item.get("ending_debt"), item.get("unsupported_funding_gap"), item.get("government_payment_arrears", item.get("contractual_arrears")), item.get("cash_balance_variance"),
    ] for item in monthly]
    _write_table(ws, 1, headers, rows, number_columns=set(range(3, 23)))
    _set_widths(ws, {1: 10, 2: 14, **{index: 19 for index in range(3, 23)}})

    ws = wb.create_sheet("Inputs and Provenance")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {1: 34, 2: 90})
    row = 1
    row = _section(ws, row, "Immutable Calculation Provenance", 2)
    provenance = [
        ("Project ID", project.id), ("Project Version ID", project_version.id), ("Project Version", project_version.version_number),
        ("Source Project Snapshot Hash", (run.input_snapshot or {}).get("source_project_snapshot_hash")),
        ("Effective Project Input Hash", sha256_json((run.input_snapshot or {}).get("project_snapshot") or {})),
        ("Policy Version ID", policy_version.id), ("Policy Version", policy_version.version_number),
        ("Policy Status", policy_version.status), ("Policy Name AR", policy_name_ar), ("Policy Name EN", policy_name_en),
        ("Policy Description AR", policy_description_ar), ("Policy Description EN", policy_description_en),
        ("Policy Hash", policy_version.snapshot_hash),
        ("Engine Version ID", engine_version.id), ("Engine Version", engine_version.engine_version), ("Adapter Version", engine_version.adapter_version),
        ("Engine Source Hash", engine_version.source_hash), ("Calculation Run ID", run.id), ("Input Hash", run.input_hash), ("Result Hash", run.result_hash),
        ("Executed At", run.completed_at.isoformat() if run.completed_at else ""),
    ]
    for label, value in provenance:
        ws.cell(row, 1, label).font = Font(color=STATIC_GRAY)
        ws.cell(row, 2, value).font = Font(color=BLACK)
        row += 1
    row += 1
    row = _section(ws, row, "Financial Model Inputs", 2)
    financial_model = (run.input_snapshot.get("project_snapshot") or {}).get("financial_model") or {}
    for key, value in _flatten(financial_model):
        ws.cell(row, 1, key).font = Font(color=INPUT_BLUE)
        ws.cell(row, 2, str(value)).font = Font(color=INPUT_BLUE)
        row += 1
    row += 1
    row = _section(ws, row, "Policy Controls", 2)
    controls = (run.input_snapshot.get("policy_snapshot") or {}).get("portal_policy") or {}
    for key, value in _flatten(controls):
        ws.cell(row, 1, key).font = Font(color="7030A0")
        ws.cell(row, 2, str(value)).font = Font(color="7030A0")
        row += 1

    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if cell.value is not None:
                    alignment = copy(cell.alignment)
                    alignment.vertical = "center"
                    cell.alignment = alignment
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _flatten(value: Any, prefix: str = ""):
    if isinstance(value, dict):
        for key in sorted(value):
            path = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten(value[key], path)
    elif isinstance(value, list):
        for index, item in enumerate(value):
            yield from _flatten(item, f"{prefix}[{index}]")
    else:
        yield prefix, value


def build_financial_pdf(
    *,
    project: Project,
    project_version: ProjectVersion,
    run: CalculationRun,
    policy_version: FinancialPolicyVersion,
    engine_version: EngineVersion,
    payload: dict[str, Any],
    language: str = "ar",
) -> bytes:
    """Create a decision-facing investment and negotiation report."""
    lang = "en" if str(language).lower().startswith("en") else "ar"
    rtl = lang == "ar"
    summary = payload.get("summary") or {}
    truth = payload.get("financial_truth") or {}
    residual = payload.get("residual_valuation") or {}
    negotiation = payload.get("negotiation_results") or []
    annual = payload.get("annual_cashflow") or []
    constraints = payload.get("constraints") or []
    audit_result = payload.get("financial_audit") or {}
    recommendation = payload.get("recommendation_validation") or {}
    model = payload.get("financial_model") or (run.input_snapshot or {}).get("project_snapshot", {}).get("financial_model") or {}
    frozen_project = (run.input_snapshot or {}).get("project_snapshot") or {}
    selected_policy_controls = policy_controls(policy_version.policy_snapshot)
    policy_name_ar = str(selected_policy_controls.get("display_name_ar") or f"السياسة v{policy_version.version_number}")
    policy_name_en = str(selected_policy_controls.get("display_name_en") or f"Policy v{policy_version.version_number}")
    policy_description_ar = str(selected_policy_controls.get("description_ar") or "")
    policy_description_en = str(selected_policy_controls.get("description_en") or "")
    currency = run.currency

    def t(ar: str, en: str) -> str:
        return en if lang == "en" else ar

    policy_name = policy_name_en if lang == "en" else policy_name_ar
    policy_description = policy_description_en if lang == "en" else policy_description_ar

    status_labels = {
        "VALIDATED": t("تم التحقق", "Validated"),
        "SUPPORTED": t("مدعومة", "Supported"),
        "CONDITIONAL": t("مشروطة", "Conditional"),
        "BLOCKED": t("محجوبة", "Blocked"),
        "PASS": t("ناجح", "Pass"),
        "FAIL": t("فشل", "Fail"),
        "VALID_RANGE": t("نطاق صالح", "Valid Range"),
        "POLICY_CAP_REACHED": t("تم بلوغ حد السياسة", "Policy Cap Reached"),
        "NOT_ESTABLISHED": t("غير مثبت", "Not Established"),
        "NO_FEASIBLE_RANGE": t("لا يوجد نطاق صالح", "No Feasible Range"),
        "INFEASIBLE": t("غير قابل للتحقق", "Infeasible"),
    }

    def status_label(value: Any) -> str:
        key = str(value or "-").strip().upper()
        return status_labels.get(key, str(value or "-"))

    def constraint_label(value: Any) -> str:
        code = str(value or "").strip().upper()
        if not code:
            return "-"
        if code == "MAX_LANDOWNER_SHARE_POLICY_CAP":
            return t("الحد الأعلى المسموح لحصة صاحب الأرض في السياسة", "Maximum Landowner Share Policy Cap")
        metadata = constraint_metadata(code)
        return str(metadata.get("title_en" if lang == "en" else "title_ar") or code)

    method_labels = {
        "GROSS_SALES": t("حصة من إجمالي المبيعات", "Gross Sales Share"),
        "NET_SALES": t("حصة من صافي المبيعات", "Net Sales Share"),
        "PROFIT_SHARE": t("مشاركة في الربح", "Profit Share"),
        "UPFRONT": t("دفعة ثابتة", "Upfront Payment"),
        "HYBRID": t("مقابل هجين", "Hybrid"),
        "MINIMUM_GUARANTEE": t("ضمان أدنى", "Minimum Guarantee"),
    }
    selected_method = str(truth.get("method") or run.selected_contract_method or (model.get("contract") or {}).get("method") or "").upper()
    selected_negotiation = next((row for row in negotiation if str(row.get("method") or "").upper() == selected_method), None)
    if selected_negotiation is None and negotiation:
        selected_negotiation = sorted(negotiation, key=lambda row: row.get("recommendation_rank") or 9999)[0]
    selected_negotiation = selected_negotiation or {}
    contract = model.get("contract") or {}

    def current_offer(row: dict[str, Any]) -> Any:
        method = str(row.get("method") or selected_method).upper()
        if method == "UPFRONT":
            return contract.get("upfront_amount")
        if method == "MINIMUM_GUARANTEE":
            return contract.get("minimum_guarantee_amount")
        return contract.get("share_rate")

    measure_type = str(selected_negotiation.get("measure_type") or truth.get("approved_measure_type") or "RATE").upper()
    offer = current_offer(selected_negotiation)
    fair_floor = selected_negotiation.get("fair_floor")
    balanced = selected_negotiation.get("balanced", selected_negotiation.get("recommended"))
    policy_ceiling = selected_negotiation.get("policy_adjusted_ceiling", selected_negotiation.get("risk_adjusted_ceiling"))
    ceiling = selected_negotiation.get("technical_ceiling")
    residual_equivalent = selected_negotiation.get("residual_equivalent_measure")

    identity = frozen_project.get("identity") or {}
    site = frozen_project.get("site") or {}
    planning = frozen_project.get("planning") or {}
    sales = model.get("sales") or {}
    delivery = model.get("delivery") or {}
    funding = model.get("funding") or {}
    finance = model.get("finance") or {}

    project_status = str(audit_result.get("validation_status") or "CONDITIONAL")
    rec_status = str(recommendation.get("status") or "CONDITIONAL")
    project_status_display = status_label(project_status)
    rec_status_display = status_label(rec_status)
    rec_reason = recommendation.get("reason_en" if lang == "en" else "reason_ar") or "-"
    compliant = t("نعم", "Yes") if payload.get("policy_compliant") else t("لا", "No")
    reconciled = t("نعم", "Yes") if payload.get("reconciliation_passed") else t("لا", "No")

    def card(label: str, value: str, kind: str = "") -> str:
        suffix = f" {escape(currency)}" if kind == "money" and value != "-" else ""
        return f"<div><strong>{escape(label)}</strong><br>{escape(value)}{suffix}</div>"

    def cards(rows: list[tuple[str, Any, str]]) -> str:
        html_parts: list[str] = []
        for label, value, kind in rows:
            if kind == "money":
                shown = _money(value)
            elif kind == "rate":
                shown = _percent(value)
            elif kind == "multiple":
                shown = _multiple(value)
            else:
                shown = str(value if value not in (None, "") else "-")
            html_parts.append(card(label, shown, kind))
        return "".join(html_parts)

    assumptions = [
        (t("تاريخ التقييم", "Valuation Date"), model.get("valuation_date") or "-"),
        (t("بداية المبيعات - شهر", "Sales Start - Month"), sales.get("start_month") or "-"),
        (t("مدة المبيعات - شهر", "Sales Duration - Months"), sales.get("duration_months") or "-"),
        (t("الخصم التجاري", "Commercial Discount"), _percent(sales.get("commercial_discount_rate"))),
        (t("مدة التنفيذ - شهر", "Construction Duration - Months"), delivery.get("construction_duration_months") or "-"),
        (t("تصاعد الكلفة", "Cost Escalation"), _percent(delivery.get("cost_escalation_rate"))),
        (t("احتياطي الكلفة", "Cost Contingency"), _percent(delivery.get("cost_contingency_rate"))),
        (t("التمويل", "Financing"), t("غير مفعل", "Off") if not finance.get("enabled") else t("مفعل", "On")),
        (t("إجمالي التزام حقوق الملكية", "Total Developer Equity Commitment"), f"{_money(funding.get('total_developer_equity'))} {currency}"),
        (t("المساهمة الافتتاحية", "Initial Equity Contribution"), f"{_money(funding.get('opening_cash'))} {currency}"),
    ]
    assumptions_rows = "".join(f"<tr><th>{escape(str(k))}</th><td>{escape(str(v))}</td></tr>" for k, v in assumptions)

    project_metrics = cards([
        (t("إجمالي المبيعات", "Gross Sales"), summary.get("gross_sales"), "money"),
        (t("صافي المبيعات", "Net Sales"), summary.get("net_sales"), "money"),
        (t("تكلفة التطوير", "Development Cost"), summary.get("development_cost"), "money"),
        (t("ربح المشروع", "Project Profit"), summary.get("project_profit"), "money"),
        (t("الربح على الكلفة", "Profit on Cost"), summary.get("project_profit_on_cost"), "rate"),
        (t("الربح على الإيراد", "Profit on Revenue"), summary.get("project_profit_on_revenue"), "rate"),
        (t("IRR المشروع", "Project IRR"), summary.get("project_irr"), "rate"),
        (t("NPV المشروع", "Project NPV"), summary.get("project_npv"), "money"),
    ])
    developer_metrics = cards([
        (t("إجمالي التزام حقوق الملكية", "Total Equity Commitment"), summary.get("total_developer_equity_commitment"), "money"),
        (t("المساهمة الافتتاحية", "Initial Equity Contribution"), summary.get("initial_equity_contribution"), "money"),
        (t("المساهمات الفعلية", "Actual Equity Contributions"), summary.get("developer_equity_contributions"), "money"),
        (t("توزيعات المطور", "Developer Distributions"), summary.get("developer_equity_distributions"), "money"),
        (t("ربح المطور", "Developer Profit"), summary.get("developer_profit"), "money"),
        (t("IRR حقوق ملكية المطور", "Developer Equity IRR"), summary.get("developer_equity_irr"), "rate"),
        (t("NPV حقوق ملكية المطور", "Developer Equity NPV"), summary.get("developer_equity_npv"), "money"),
        ("MOIC", summary.get("developer_equity_multiple"), "multiple"),
        (t("ذروة حقوق الملكية المطلوبة", "Peak Equity Requirement"), summary.get("peak_equity"), "money"),
    ])
    landowner_metrics = cards([
        (t("الآلية المختارة", "Selected Mechanism"), method_labels.get(selected_method, selected_method or "-"), "text"),
        (t("المقابل الاسمي", "Nominal Consideration"), summary.get("government_consideration"), "money"),
        (t("القيمة الحالية لصاحب الأرض", "Landowner NPV"), summary.get("government_consideration_npv"), "money"),
        (t("المتحصل النقدي الفعلي", "Actual Cash Receipts"), truth.get("landowner_cash_receipts", summary.get("landowner_cash_receipts")), "money"),
    ])
    residual_metrics = cards([
        (t("القيمة المتبقية التطويرية", "Residual Land Value"), residual.get("residual_land_value"), "money"),
        (t("قدرة الأرض بطريقة DCF", "Land Capacity DCF"), residual.get("land_capacity_dcf"), "money"),
        (t("GDV قبل الأرض", "Pre-Land GDV"), residual.get("gross_development_value"), "money"),
        (t("ربح المطور المستهدف على الكلفة", "Target Developer Profit on Cost"), residual.get("target_developer_profit_on_cost"), "rate"),
    ])

    negotiation_rows = "".join(
        "<tr>"
        f"<td>{escape(method_labels.get(str(row.get('method') or '').upper(), str(row.get('method') or '-')))}</td>"
        f"<td>{escape(status_label(row.get('status')))}</td>"
        f"<td>{escape(_value_by_type(row.get('fair_floor'), row.get('measure_type')) if row.get('fair_floor') not in (None, '') else t('غير مثبت', 'Not established'))}</td>"
        f"<td>{escape(_value_by_type(row.get('balanced', row.get('recommended')), row.get('measure_type')))}</td>"
        f"<td>{escape(_value_by_type(row.get('policy_adjusted_ceiling', row.get('risk_adjusted_ceiling')), row.get('measure_type')))}</td>"
        f"<td>{escape(_value_by_type(row.get('residual_equivalent_measure'), row.get('measure_type')) if row.get('residual_equivalent_measure') not in (None, '') else '-')}</td>"
        f"<td>{escape(_value_by_type(row.get('technical_ceiling'), row.get('measure_type')))}</td>"
        f"<td>{escape(constraint_label(row.get('governing_constraint_id')))}</td>"
        "</tr>" for row in negotiation
    )
    annual_rows = "".join(
        "<tr>"
        f"<td>{escape(str(row.get('year') or '-'))}</td>"
        f"<td>{escape(_money(row.get('gross_collections')))}</td>"
        f"<td>{escape(_money(row.get('actual_cost')))}</td>"
        f"<td>{escape(_money(row.get('equity_contribution')))}</td>"
        f"<td>{escape(_money(row.get('financing_draw')))}</td>"
        f"<td>{escape(_money(row.get('landowner_cash_receipt', row.get('government_payment'))))}</td>"
        f"<td>{escape(_money(row.get('developer_distribution')))}</td>"
        f"<td>{escape(_money(row.get('ending_cash')))}</td>"
        f"<td>{escape(_money(row.get('ending_debt')))}</td>"
        "</tr>" for row in annual
    )
    failed_constraints = [row for row in constraints if row.get("passed") is False]
    technical_rows = "".join(
        "<tr>"
        f"<td>{escape(str(row.get('constraint_id') or '-'))}</td>"
        f"<td>{escape(str((row.get('title_en') if lang == 'en' else row.get('title_ar')) or constraint_label(row.get('constraint_id'))))}</td>"
        f"<td>{escape(_constraint_value(row.get('constraint_id'), row.get('actual')))}</td>"
        f"<td>{escape(str(row.get('operator') or ''))} {escape(_constraint_value(row.get('constraint_id'), row.get('threshold')))}</td>"
        f"<td>{escape(status_label('PASS' if row.get('passed') else 'FAIL'))}</td>"
        "</tr>" for row in (failed_constraints or constraints)
    )

    floor_display = _value_by_type(fair_floor, measure_type) if fair_floor not in (None, "") else t("غير مثبت", "Not established")
    balanced_display = _value_by_type(balanced, measure_type)
    policy_ceiling_display = _value_by_type(policy_ceiling, measure_type)
    ceiling_display = _value_by_type(ceiling, measure_type)
    residual_equivalent_display = _value_by_type(residual_equivalent, measure_type) if residual_equivalent not in (None, "") else "-"
    offer_display = _value_by_type(offer, measure_type)
    range_reason = selected_negotiation.get("fair_floor_reason_en" if lang == "en" else "fair_floor_reason_ar") or ""
    floor_raw = "" if fair_floor in (None, "") else str(fair_floor)
    unit = currency if measure_type == "AMOUNT" else ""

    explanation_rows = {str(item.get("code") or ""): item for item in (selected_negotiation.get("negotiation_explanations") or [])}

    def explanation_detail(code: str) -> str:
        item = explanation_rows.get(code) or {}
        return str(item.get("detail_en" if lang == "en" else "detail_ar") or item.get("detail_ar" if lang == "en" else "detail_en") or "")

    minimum_case = selected_negotiation.get("minimum_case") or {}
    balanced_case = selected_negotiation.get("balanced_case") or selected_negotiation.get("recommended_case") or {}
    policy_case = selected_negotiation.get("policy_adjusted_ceiling_case") or selected_negotiation.get("risk_adjusted_ceiling_case") or {}
    ceiling_case = selected_negotiation.get("ceiling_case") or {}
    offer_case = selected_negotiation.get("offer_case") or {}
    floor_explanation = explanation_rows.get("PUBLIC_VALUE_FLOOR") or {}
    risk_explanation = explanation_rows.get("RISK_ADJUSTMENT") or {}
    technical_explanation = explanation_rows.get("TECHNICAL_CEILING") or {}
    offer_explanation = explanation_rows.get("OFFER_POSITION") or {}
    governing = technical_explanation.get("governing_constraint") or {}
    eligible_base = selected_negotiation.get("eligible_base_total")
    basis_label = str(selected_negotiation.get("basis_label_ar" if lang == "ar" else "basis_label_en") or selected_negotiation.get("basis_label") or t("وعاء العقد المؤهل", "Eligible contract base"))

    minimum_body = explanation_detail("PUBLIC_VALUE_FLOOR") or range_reason or t(
        "أقل مقابل يحقق الحد الأدنى المطلوب لقيمة صاحب الأرض مع بقاء جميع شروط الجدوى ناجحة.",
        "The lowest consideration satisfying the required landowner value while all feasibility constraints remain satisfied.",
    )
    minimum_evidence = "" if fair_floor in (None, "") else (
        f"{t('NPV صاحب الأرض', 'Landowner NPV')}: {_money(minimum_case.get('government_gross_npv', minimum_case.get('government_npv', floor_explanation.get('actual_public_npv'))))} {currency} · "
        f"{t('الحد المطلوب', 'Required')}: {_money(floor_explanation.get('required_public_npv'))} {currency}"
    )
    balanced_body = explanation_detail("BALANCED_POINT") or t(
        "نقطة توصية داخل المجال المتحفظ تحددها نسخة السياسة المختارة، ولا تساوي السقف الفني تلقائياً.",
        "A recommended point inside the policy-adjusted range, determined by the selected policy version and not automatically equal to the technical ceiling.",
    )
    balanced_evidence = (
        f"{t('IRR المطور', 'Developer IRR')}: {_percent(balanced_case.get('developer_equity_irr', balanced_case.get('developer_irr')))} · "
        f"{t('NPV صاحب الأرض', 'Landowner NPV')}: {_money(balanced_case.get('government_gross_npv', balanced_case.get('government_npv')))} {currency} · "
        f"{t('فجوة التمويل', 'Funding Gap')}: {_money(balanced_case.get('peak_funding_gap'))} {currency}"
    )
    residual_body = t(
        "تحويل القيمة المتبقية التطويرية إلى قيمة مكافئة على وعاء الآلية المختارة لسهولة المقارنة. وهي مؤشر قدرة تطويرية وليست تقييماً سوقياً مستقلاً.",
        "Conversion of development residual land value to an equivalent measure on the selected mechanism basis. It is a development-capacity indicator, not an independent market valuation.",
    )
    residual_evidence = ""
    if residual.get("residual_land_value") not in (None, ""):
        if measure_type == "RATE" and eligible_base not in (None, "", 0, "0"):
            residual_evidence = f"{_money(residual.get('residual_land_value'))} {currency} ÷ {basis_label} {_money(eligible_base)} {currency} = {residual_equivalent_display}"
        else:
            residual_evidence = f"{t('القيمة المتبقية التطويرية', 'Residual Land Value')}: {_money(residual.get('residual_land_value'))} {currency}"
    policy_body = explanation_detail("RISK_ADJUSTMENT") or t(
        "حد تفاوضي متحفظ داخل القدرة الفنية، تحدده معاملات التحفظ وهامش أمان المطور في نسخة السياسة.",
        "A conservative negotiating limit within technical capacity, set by policy conservatism and developer safety factors.",
    )
    policy_evidence = (
        f"{t('IRR المطور', 'Developer IRR')}: {_percent(policy_case.get('developer_equity_irr', policy_case.get('developer_irr')))} · "
        f"{t('NPV صاحب الأرض', 'Landowner NPV')}: {_money(policy_case.get('government_gross_npv', policy_case.get('government_npv')))} {currency}"
    )
    if risk_explanation.get("capacity_factor") not in (None, ""):
        policy_evidence += f" · {t('معامل القدرة المطبق', 'Applied capacity factor')}: {_percent(risk_explanation.get('capacity_factor'))}"
    technical_body = explanation_detail("TECHNICAL_CEILING") or t(
        "أعلى مقابل يظل عنده المشروع مستوفياً لجميع القيود قبل أول نقطة فشل. وهو حد قدرة وليس توصية تفاوضية.",
        "The highest consideration at which all constraints remain satisfied before the first failing point. It is a capacity limit, not a negotiation recommendation.",
    )
    if governing:
        technical_evidence = (
            f"{constraint_label(governing.get('id'))}: {_constraint_value(governing.get('id'), governing.get('actual'))} "
            f"{governing.get('operator') or ''} {_constraint_value(governing.get('id'), governing.get('threshold'))}"
        )
    else:
        technical_evidence = f"{t('القيد الحاكم', 'Governing constraint')}: {constraint_label(selected_negotiation.get('governing_constraint_id'))} · {t('IRR المطور', 'Developer IRR')}: {_percent(ceiling_case.get('developer_equity_irr', ceiling_case.get('developer_irr')))}"
    offer_body = explanation_detail("OFFER_POSITION") or t(
        "القيمة المدخلة حالياً في العقد، ويقارنها النظام بالحدود المالية والتفاوضية على الوعاء نفسه.",
        "The currently entered contract value, compared against financial and negotiation boundaries on the same calculation basis.",
    )
    offer_evidence = (
        f"{t('المقابل الاسمي', 'Nominal consideration')}: {_money(offer_case.get('government_value'))} {currency} · "
        f"{t('NPV صاحب الأرض', 'Landowner NPV')}: {_money(offer_case.get('government_gross_npv', offer_case.get('government_npv')))} {currency} · "
        f"{t('IRR المطور', 'Developer IRR')}: {_percent(offer_case.get('developer_equity_irr', offer_case.get('developer_irr')))}"
    )

    def explanation_card(tone: str, title: str, value: str, body: str, evidence: str) -> str:
        return (
            f'<div class="explanation-card" data-tone="{escape(tone)}" data-title="{escape(title)}" '
            f'data-value="{escape(value)}" data-body="{escape(body)}" data-evidence="{escape(evidence)}">'
            f'<strong>{escape(title)}</strong><b>{escape(value)}</b><p>{escape(body)}</p><small>{escape(evidence)}</small></div>'
        )

    negotiation_explanation_cards = "".join([
        explanation_card("minimum", t("الحد الأدنى المقبول", "Minimum Acceptable"), floor_display, minimum_body, minimum_evidence),
        explanation_card("balanced", t("النقطة المتوازنة", "Balanced Point"), balanced_display, balanced_body, balanced_evidence),
        explanation_card("residual", t("القيمة المتبقية المكافئة", "Residual Equivalent"), residual_equivalent_display, residual_body, residual_evidence),
        explanation_card("policy", t("السقف المتحفظ وفق السياسة", "Policy-Adjusted Ceiling"), policy_ceiling_display, policy_body, policy_evidence),
        explanation_card("technical", t("السقف الفني", "Technical Ceiling"), ceiling_display, technical_body, technical_evidence),
        explanation_card("offer", t("العرض الحالي", "Current Offer"), offer_display, offer_body, offer_evidence),
    ])
    negotiation_summary = str(selected_negotiation.get("negotiation_summary_en" if lang == "en" else "negotiation_summary_ar") or "")
    negotiation_reading = " ".join(part for part in (negotiation_summary, offer_body) if part).strip()
    axis_values = [value for value in (fair_floor, balanced, policy_ceiling, residual_equivalent, ceiling, offer) if value not in (None, "")]
    try:
        axis_max_value = max(axis_values, key=lambda value: float(value)) if axis_values else offer
    except (TypeError, ValueError):
        axis_max_value = offer
    axis_max_display = _value_by_type(axis_max_value, measure_type)

    executive_callout = f"{t('حالة التدقيق المالي', 'Financial validation')}: {project_status_display}. {t('حالة التوصية', 'Recommendation status')}: {rec_status_display}. {rec_reason}"
    land_explanation = t(
        "مؤشرات القيمة المتبقية تقيس قدرة المشروع التطويرية على تحمل قيمة الأرض قبل إدخال مقابل الأرض. لا تمثل تقييماً سوقياً مستقلاً، ولا تستبدل نطاق التفاوض الناتج عن قيود عائد المطور وقيمة صاحب الأرض.",
        "Residual indicators measure development capacity to support land value before land consideration. They are not an independent market valuation and do not replace the negotiation range produced by developer-return and landowner-value constraints.",
    )
    floor_note = range_reason or t("لم يتم تثبيت حد عادل اقتصادي لهذه الآلية.", "A minimum acceptable economic consideration has not been established for this mechanism; the system therefore does not present a near-zero value as a defensible floor.")
    doc_title = t("التقرير المالي والاستثماري", "Financial & Investment Report")
    direction = "rtl" if rtl else "ltr"
    glossary = [
        (t("إجمالي المبيعات", "Gross Sales"), t("القيمة الاسمية للمبيعات قبل الحسميات والحوافز والمرتجعات.", "Nominal sales value before discounts, incentives and refunds.")),
        (t("صافي المبيعات", "Net Sales"), t("المبيعات بعد الحسميات والحوافز والمرتجعات المؤهلة وفق الافتراضات المعتمدة.", "Sales after eligible discounts, incentives and refunds under approved assumptions.")),
        (t("IRR المشروع", "Project IRR"), t("العائد السنوي على التدفق النقدي غير الممول للمشروع قبل أثر تمويل المطور.", "Annual return on project unlevered cash flow before developer-financing effects.")),
        (t("NPV المشروع", "Project NPV"), t("القيمة الحالية لتدفق المشروع عند معدل الخصم المعتمد.", "Present value of project cash flow at the approved discount rate.")),
        (t("IRR حقوق ملكية المطور", "Developer Equity IRR"), t("العائد السنوي على الأموال التي يضخها المطور فعلياً، مع مراعاة توقيت المساهمات والتوزيعات.", "Annual return on actual developer equity, considering contribution and distribution timing.")),
        (t("NPV حقوق ملكية المطور", "Developer Equity NPV"), t("القيمة الحالية لتدفقات المطور عند معدل الخصم المعتمد.", "Present value of developer cash flows at the approved discount rate.")),
        ("MOIC", t("إجمالي توزيعات المطور مقسوماً على حقوق الملكية المضخوخة؛ لا يقيس توقيت التدفقات.", "Developer distributions divided by contributed equity; it does not measure timing.")),
        (t("ذروة حقوق الملكية المطلوبة", "Peak Equity Requirement"), t("أعلى رصيد تراكمي من حقوق ملكية المطور يكون معرضاً داخل المشروع في أي شهر.", "Highest cumulative developer equity exposed in the project in any month.")),
        (t("القيمة الحالية لصاحب الأرض", "Landowner NPV"), t("القيمة الحالية لجميع المقبوضات المتوقعة لصاحب الأرض مع مراعاة توقيتها.", "Present value of all expected landowner receipts, considering timing.")),
        (t("الحد الأدنى المقبول", "Minimum Acceptable"), t("أدنى مقابل يحقق قيمة الاسترداد الدنيا المعتمدة لصاحب الأرض؛ يظهر غير مثبت عند غياب حد اقتصادي مادي.", "Minimum consideration satisfying approved landowner recovery; shown as not established when no material economic floor exists.")),
        (t("النقطة المتوازنة", "Balanced Point"), t("نقطة داخل المجال المتحفظ تحددها نسبة موضع معتمدة في السياسة بين الحد الأدنى المقبول والسقف المتحفظ؛ لا تساوي السقف الفني تلقائياً.", "A point inside the policy-adjusted range, positioned by the approved policy factor between Minimum Acceptable and the policy-adjusted ceiling; it does not automatically equal the technical ceiling.")),
        (t("السقف المتحفظ وفق السياسة", "Policy-Adjusted Ceiling"), t("حد تفاوضي متحفظ داخل القدرة الفنية، تحدده عوامل السعة والتحفظ المعتمدة في نسخة السياسة.", "A conservative negotiation limit within technical capacity, determined by the capacity and conservatism factors of the selected policy version.")),
        (t("السقف الفني", "Technical Ceiling"), t("أعلى مقابل يثبت عنده بقاء جميع شروط الجدوى والسيولة والإقفال ناجحة قبل أول نقطة فشل؛ وإذا توقف البحث عند حد إداري يظهر ذلك كحد بحث لا كسقف اقتصادي مثبت.", "Highest consideration at which all feasibility, liquidity and close constraints remain satisfied before the first failing point; an administrative search limit is disclosed as such, not as an established economic ceiling.")),
        (t("القيمة المتبقية التطويرية", "Residual Land Value"), t("قدرة نظرية للمشروع على تحمل قيمة الأرض بعد كلف التطوير والعائد المستهدف؛ ليست تقييماً سوقياً مستقلاً.", "Theoretical land-paying capacity after development costs and target return; not an independent market valuation.")),
        (t("المصالحة النقدية", "Cash Reconciliation"), t("تحقق شهري من تساوي النقد الافتتاحي ومصادر التمويل مع الاستخدامات والنقد الختامي.", "Monthly check that opening cash and funding sources equal cash uses and ending cash.")),
    ]
    glossary_rows = "".join(
        f"<tr><th>{escape(term)}</th><td>{escape(definition)}</td></tr>" for term, definition in glossary
    )

    html = f'''<!doctype html><html lang="{lang}" dir="{direction}"><head><meta charset="utf-8"><title>LandValue360 {escape(doc_title)} - {escape(project.reference)}</title></head><body>
    <header class="cover page">
      <h1>{escape(doc_title)}</h1><h2>{escape(project.name)}</h2>
      <p>{escape(project.reference)} | {escape(t('إصدار المشروع', 'Project Version'))} {project_version.version_number} | {escape(t('السياسة المالية', 'Financial Policy'))}: {escape(policy_name)} v{policy_version.version_number} | {escape(t('العملة', 'Currency'))}: {escape(currency)}</p>
      <div class="metric-grid">{card(t('التدقيق المالي','Financial Validation'), project_status_display)}{card(t('التوصية','Recommendation'), rec_status_display)}{card(t('مطابقة السياسة','Policy Compliance'), compliant)}{card(t('المصالحة الشهرية','Monthly Reconciliation'), reconciled)}</div>
      <div class="cover-disclaimer">{escape(t('هذا التقرير مبني على نموذج مالي شهري واحد وعلى إصدار ثابت من المشروع والسياسة والمحرك. التوصية لا تعتبر مدعومة إذا فشل التدقيق المالي أو قيود الإقفال.', 'This report is based on one monthly financial model and immutable project, policy, and engine versions. A recommendation is not treated as supported when financial validation or closing constraints fail.'))}</div>
    </header>

    <section class="page"><h2>{escape(t('1. الملخص التنفيذي', '1. Executive Summary'))}</h2><div class="callout">{escape(executive_callout)}</div>
      <div class="metric-grid">{card(t('ربح المشروع','Project Profit'), _money(summary.get('project_profit')), 'money')}{card(t('IRR حقوق ملكية المطور','Developer Equity IRR'), _percent(summary.get('developer_equity_irr')))}{card(t('NPV حقوق ملكية المطور','Developer Equity NPV'), _money(summary.get('developer_equity_npv')), 'money')}{card('MOIC', _multiple(summary.get('developer_equity_multiple')))}{card(t('NPV صاحب الأرض','Landowner NPV'), _money(summary.get('government_consideration_npv')), 'money')}{card(t('الآلية المختارة','Selected Mechanism'), method_labels.get(selected_method, selected_method or '-'))}</div>
      <p>{escape(t('تعرض النتائج بثلاث طبقات مستقلة: اقتصاديات المشروع، اقتصاديات المطور، واقتصاديات صاحب الأرض، ثم يربط بينها تحليل التفاوض.', 'Results are presented in three separate layers: project economics, developer economics, and landowner economics, then connected through the negotiation analysis.'))}</p>
    </section>

    <section class="page"><h2>{escape(t('2. معلومات المشروع والافتراضات', '2. Project Information & Assumptions'))}</h2>
      <table><tr><th>{escape(t('المشروع','Project'))}</th><td>{escape(project.name)}</td></tr><tr><th>{escape(t('المرجع','Reference'))}</th><td>{escape(project.reference)}</td></tr><tr><th>{escape(t('الموقع','Location'))}</th><td>{escape(str(site.get('location') or identity.get('location') or '-'))}</td></tr><tr><th>{escape(t('مدة المشروع','Project Duration'))}</th><td>{escape(str(planning.get('project_duration_months') or summary.get('original_project_duration_months') or '-'))}</td></tr><tr><th>{escape(t('السياسة المالية المختارة','Selected Financial Policy'))}</th><td>{escape(policy_name)} - v{policy_version.version_number}</td></tr><tr><th>{escape(t('وصف السياسة','Policy Description'))}</th><td>{escape(policy_description or '-')}</td></tr>{assumptions_rows}</table>
      <div class="callout">{escape(t('للمستخدم العادي، منحنيات المبيعات والكلف وافتراضات التحصيل والتمويل المتقدمة تأتي من نسخة السياسة المالية المعتمدة. يمكن للمحلل المخول فقط تفعيل افتراضات خاصة بالمشروع.', 'For standard users, sales/cost curves, collections, and advanced financing assumptions are governed by the approved financial policy. Only an authorized analyst may enable project-specific overrides.'))}</div>
    </section>

    <section class="page"><h2>{escape(t('3. اقتصاديات المشروع', '3. Project Economics'))}</h2><div class="metric-grid">{project_metrics}</div>
      <table><tr><th>{escape(t('المدة الأصلية','Original Duration'))}</th><th>{escape(t('المدة المعدلة','Adjusted Duration'))}</th><th>{escape(t('ذروة فجوة التمويل','Peak Funding Gap'))}</th><th>{escape(t('الرصيد الختامي','Ending Cash'))}</th></tr><tr><td>{escape(str(summary.get('original_project_duration_months') or '-'))}</td><td>{escape(str(summary.get('adjusted_project_duration_months') or '-'))}</td><td>{escape(_money(summary.get('peak_funding_gap')))} {escape(currency)}</td><td>{escape(_money(summary.get('ending_cash')))} {escape(currency)}</td></tr></table>
    </section>

    <section class="page"><h2>{escape(t('4. اقتصاديات المطور', '4. Developer Economics'))}</h2><div class="metric-grid">{developer_metrics}</div>
      <p>{escape(t('إجمالي التزام حقوق الملكية يشمل المساهمة الافتتاحية؛ لا تتم إضافة الرصيد الافتتاحي مرة ثانية فوق إجمالي الالتزام.', 'Total developer equity commitment includes the initial contribution; opening equity is not added a second time on top of the total commitment.'))}</p>
      <table><tr><th>{escape(t('ذروة الدين','Peak Debt'))}</th><th>{escape(t('الفائدة','Interest'))}</th><th>{escape(t('رسوم التمويل','Financing Fees'))}</th><th>{escape(t('الدين الختامي','Terminal Debt'))}</th></tr><tr><td>{escape(_money(summary.get('peak_debt')))} {escape(currency)}</td><td>{escape(_money(summary.get('interest_total')))} {escape(currency)}</td><td>{escape(_money(summary.get('financing_fees_total')))} {escape(currency)}</td><td>{escape(_money(summary.get('terminal_debt')))} {escape(currency)}</td></tr></table>
    </section>

    <section class="page"><h2>{escape(t('5. اقتصاديات صاحب الأرض', '5. Landowner Economics'))}</h2><div class="metric-grid">{landowner_metrics}</div>
      <table><tr><th>{escape(t('الحد الأدنى المقبول','Minimum Acceptable'))}</th><th>{escape(t('النقطة المتوازنة','Balanced'))}</th><th>{escape(t('السقف المتحفظ','Policy-Adjusted Ceiling'))}</th><th>{escape(t('المعادِل للقيمة المتبقية','Residual Equivalent'))}</th><th>{escape(t('العرض الحالي','Current Offer'))}</th><th>{escape(t('السقف الفني / حد البحث','Technical Ceiling / Search Cap'))}</th></tr><tr><td>{escape(floor_display)}</td><td>{escape(balanced_display)}</td><td>{escape(policy_ceiling_display)}</td><td>{escape(residual_equivalent_display)}</td><td>{escape(offer_display)}</td><td>{escape(ceiling_display)}</td></tr></table><div class="callout">{escape(floor_note)}</div>
    </section>

    <section class="page"><h2>{escape(t('6. تقييم قدرة الأرض والمجال التفاوضي', '6. Land Capacity & Negotiation Range'))}</h2><div class="metric-grid">{residual_metrics}</div><div class="callout">{escape(land_explanation)}</div>
      <div class="callout decision-box">{escape(negotiation_reading or executive_callout)}</div>
      <div class="negotiation-band" data-title="{escape(method_labels.get(str(selected_negotiation.get('method') or selected_method).upper(), selected_method or '-'))}" data-floor="{escape(floor_raw)}" data-balanced="{escape(str(balanced or ''))}" data-policy="{escape(str(policy_ceiling or ''))}" data-residual="{escape(str(residual_equivalent or ''))}" data-ceiling="{escape(str(ceiling or ''))}" data-offer="{escape(str(offer or ''))}" data-floor-display="{escape(floor_display)}" data-balanced-display="{escape(balanced_display)}" data-policy-display="{escape(policy_ceiling_display)}" data-residual-display="{escape(residual_equivalent_display)}" data-ceiling-display="{escape(ceiling_display)}" data-offer-display="{escape(offer_display)}" data-axis-max-display="{escape(axis_max_display)}" data-unit="{escape(unit)}" data-floor-label="{escape(t('الحد الأدنى المقبول','Minimum Acceptable'))}" data-balanced-label="{escape(t('متوازن','Balanced'))}" data-policy-label="{escape(t('السقف المتحفظ','Policy Ceiling'))}" data-residual-label="{escape(t('القيمة المتبقية','Residual'))}" data-ceiling-label="{escape(t('السقف الفني','Technical Ceiling'))}" data-offer-label="{escape(t('العرض الحالي','Current Offer'))}"></div>
      <h3>{escape(t('كيف تم تحديد هذا النطاق؟','How Was This Range Determined?'))}</h3><div class="explanation-grid">{negotiation_explanation_cards}</div>
      <h3>{escape(t('مقارنة آليات التعاقد','Contract Mechanism Comparison'))}</h3><table class="compact"><tr><th>{escape(t('الآلية','Mechanism'))}</th><th>{escape(t('الحالة','Status'))}</th><th>{escape(t('الحد الأدنى المقبول','Minimum Acceptable'))}</th><th>{escape(t('متوازن','Balanced'))}</th><th>{escape(t('السقف المتحفظ','Policy Ceiling'))}</th><th>{escape(t('معادل القيمة المتبقية','Residual Equivalent'))}</th><th>{escape(t('السقف الفني','Technical Ceiling'))}</th><th>{escape(t('القيد الحاكم','Governing Constraint'))}</th></tr>{negotiation_rows}</table>
    </section>

    <section class="page"><h2>{escape(t('7. التدفق النقدي السنوي', '7. Annual Cash Flow'))}</h2><table class="compact"><tr><th>{escape(t('السنة','Year'))}</th><th>{escape(t('التحصيلات','Collections'))}</th><th>{escape(t('الكلف المدفوعة','Paid Costs'))}</th><th>{escape(t('حقوق الملكية','Equity'))}</th><th>{escape(t('سحب التمويل','Debt Draw'))}</th><th>{escape(t('صاحب الأرض','Landowner'))}</th><th>{escape(t('توزيعات المطور','Developer Distributions'))}</th><th>{escape(t('النقد الختامي','Ending Cash'))}</th><th>{escape(t('الدين الختامي','Ending Debt'))}</th></tr>{annual_rows}</table></section>

    <section class="page"><h2>{escape(t('8. المراجعة الفنية', '8. Technical Validation'))}</h2><div class="metric-grid">{card(t('حالة التدقيق','Audit Status'), project_status_display)}{card(t('مطابقة السياسة','Policy Compliance'), compliant)}{card(t('المصالحة النقدية','Cash Reconciliation'), reconciled)}{card(t('حالة التوصية','Recommendation Status'), rec_status_display)}</div><table class="compact"><tr><th>ID</th><th>{escape(t('القيد','Constraint'))}</th><th>{escape(t('القيمة الفعلية','Actual'))}</th><th>{escape(t('الحد','Limit'))}</th><th>{escape(t('النتيجة','Result'))}</th></tr>{technical_rows}</table></section>

    <section class="page"><h2>{escape(t('9. دليل المؤشرات المالية', '9. Financial Indicator Guide'))}</h2><p>{escape(t('تشرح هذه الصفحة المصطلحات الأساسية المستخدمة في التقرير بلغة قرار مبسطة. التعاريف لا تغير المعادلات أو نتائج التشغيل.', 'This page explains the report’s core terms in decision-oriented language. Definitions do not alter formulas or calculation results.'))}</p><table class="compact"><tr><th>{escape(t('المؤشر','Indicator'))}</th><th>{escape(t('المعنى','Meaning'))}</th></tr>{glossary_rows}</table></section>

    <section class="page"><h2>{escape(t('ملحق - المراجع الفنية والإصدارات', 'Appendix - Technical Provenance'))}</h2><table><tr><th>{escape(t('معرف التشغيل الحسابي','Calculation Run ID'))}</th><td>{escape(run.id)}</td></tr><tr><th>{escape(t('إصدار المشروع','Project Version'))}</th><td>v{project_version.version_number} / {escape(project_version.id)}</td></tr><tr><th>{escape(t('إصدار السياسة','Policy Version'))}</th><td>{escape(policy_name)} - v{policy_version.version_number} / {escape(policy_version.id)} / {escape(policy_version.status)}</td></tr><tr><th>{escape(t('وصف السياسة','Policy Description'))}</th><td>{escape(policy_description or '-')}</td></tr><tr><th>{escape(t('المحرك والمحول','Engine and Adapter'))}</th><td>{escape(engine_version.engine_version)} / {escape(t('المحول','Adapter'))} {escape(engine_version.adapter_version)}</td></tr><tr><th>{escape(t('بصمة المدخلات','Input Hash'))}</th><td>{escape(run.input_hash)}</td></tr><tr><th>{escape(t('بصمة النتائج','Result Hash'))}</th><td>{escape(run.result_hash or '')}</td></tr></table><p>{escape(t('القيمة المتبقية التطويرية ليست تقييماً سوقياً مستقلاً. يلزم التعامل مع الافتراضات القانونية والضريبية والسوقية التي تقع خارج نطاق هذا المحرك بشكل مستقل قبل الالتزام التعاقدي.', 'The development residual indication is not an independent market valuation. Legal, tax, and market assumptions outside this engine must be addressed separately before contractual commitment.'))}</p></section>
    </body></html>'''
    return render_html_pdf(html)

