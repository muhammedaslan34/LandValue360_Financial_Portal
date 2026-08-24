from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from landvalue360_portal.database import session_scope
from landvalue360_portal.models import (
    CalculationRun,
    CalculationRunResult,
    EngineVersion,
    FinancialPolicyVersion,
    MonthlyCashflowSnapshot,
    NegotiationResult,
    Organization,
    Project,
    ProjectVersion,
)
from landvalue360_portal.financial_engine import (
    apply_policy_controls,
    default_financial_policy_snapshot,
    normalize_financial_model,
    sha256_json,
)
from landvalue360_portal.services import create_staff_user


def register(client, email='finance-owner@example.com', org='Finance Owner'):
    response = client.post('/api/auth/register', json={
        'email': email,
        'password': 'StrongPass123!',
        'full_name': 'Finance Owner',
        'organization_name': org,
        'country': 'SY',
        'phone': '',
        'accepted_terms': True,
    })
    assert response.status_code == 200, response.text
    return response.json()['csrf_token']


def login(client, email):
    response = client.post('/api/auth/login', json={'email': email, 'password': 'StrongPass123!'})
    assert response.status_code == 200, response.text
    return response.json()['csrf_token']


def project_payload():
    return {
        'name': 'Financial Portal Golden Case',
        'description': 'Authoritative monthly model smoke case',
        'currency': 'USD',
        'gross_land_area_sqm': '10000',
        'excluded_land_area_sqm': '0',
        'title_reference': 'FIN-001',
        'location': 'Damascus',
        'current_land_value': '2500000',
        'far': '2',
        'bcr': '0.4',
        'planning_status': 'Concept',
        'project_duration_months': 36,
        'sales_duration_months': 30,
        'land_uses': [{'code': 'INVESTMENT', 'name': 'Investment', 'percentage': '100'}],
        'products': [{
            'code': 'RES', 'name': 'Residential', 'allocation_percentage': '100',
            'sellable_efficiency_percentage': '80', 'unit_selling_price': '1000',
            'currency': 'USD', 'price_source': 'Golden case', 'evidence_confidence': 'HIGH',
        }],
        'costs': [{
            'name': 'Construction', 'category': 'CONSTRUCTION', 'amount': '8000000',
            'currency': 'USD', 'quantity_basis': None, 'quantity': None, 'unit_cost': None,
            'developer_share_percentage': '100', 'net_sales_deductible': True,
            'notes': None, 'source': 'Golden case', 'evidence_confidence': 'HIGH',
        }],
    }


def setup_project(client):
    csrf = register(client)
    with session_scope() as db:
        org = db.scalar(select(Organization).where(Organization.slug == 'finance-owner'))
    response = client.post('/api/projects', json={
        'organization_id': org.id, 'name': 'Financial Portal Golden Case', 'currency': 'USD',
    }, headers={'X-CSRF-Token': csrf})
    assert response.status_code == 201, response.text
    project_id = response.json()['id']
    response = client.put(f'/api/projects/{project_id}', json=project_payload(), headers={'X-CSRF-Token': csrf})
    assert response.status_code == 200, response.text
    return project_id, csrf


def test_authoritative_financial_run_and_reports(client):
    project_id, csrf = setup_project(client)
    state = client.get(f'/api/projects/{project_id}/financial')
    assert state.status_code == 200, state.text
    payload = state.json()
    assert payload['project']['currency'] == 'USD'
    assert payload['engine']['engine_version'] == '2.1.1'
    assert payload['policy']['version_number'] == 1
    assert payload['advanced_financial_access'] is False
    assert 'finance' not in payload['financial_model']
    assert 'curve_type' not in payload['financial_model']['sales']
    assert 'construction_curve_type' not in payload['financial_model']['delivery']
    assert 'advanced_defaults' not in payload['policy']['controls']

    model = payload['financial_model']
    model['sales']['duration_months'] = 30
    model['sales']['collection_rules'] = [
        {'lag_months': 0, 'weight': '0.30', 'label': 'Contract'},
        {'lag_months': 12, 'weight': '0.40', 'label': 'Construction'},
        {'lag_months': 24, 'weight': '0.30', 'label': 'Handover'},
    ]
    model['funding']['opening_cash'] = '12000000'
    model['funding']['total_developer_equity'] = '12000000'
    model['funding']['committed_financing'] = '99999999'
    model['sales']['curve_type'] = 'LINEAR'
    model['sales']['collection_rules'] = [{'lag_months': 0, 'weight': '1', 'label': 'Crafted'}]
    model['finance'] = {'enabled': True, 'allow_negative_cash': True, 'annual_interest_rate': '0.99'}
    response = client.put(
        f'/api/projects/{project_id}/financial', json=model, headers={'X-CSRF-Token': csrf},
    )
    assert response.status_code == 200, response.text
    saved_model = response.json()['financial_model']
    assert 'finance' not in saved_model
    assert 'curve_type' not in saved_model['sales']
    assert 'collection_rules' not in saved_model['sales']
    assert 'committed_financing' not in saved_model['funding']
    assert saved_model['funding']['total_developer_equity'] == '12000000'
    assert saved_model['funding']['committed_additional_equity'] == '0'

    response = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={'project_version_id': payload['project_version']['id']},
        headers={'X-CSRF-Token': csrf},
    )
    assert response.status_code == 201, response.text
    run = response.json()
    assert run['status'] == 'COMPLETED'
    assert len(run['input_hash']) == 64 and len(run['result_hash']) == 64
    assert run['project_version_number'] == 1
    assert run['financial_policy_version_number'] == 1
    assert run['engine_version_label'] == '2.1.1'
    assert run['financial_truth']['cash_reconciliation_passed'] is True
    assert run['residual_valuation']['calculated_before_landowner_consideration'] is True
    assert run['residual_valuation']['indication_type'] == 'DEVELOPMENT_RESIDUAL_INDICATION'
    assert len(run['monthly_cashflow']) >= 36
    assert len(run['annual_cashflow']) >= 3
    assert run['financial_audit']['validation_status'] in {'VALIDATED', 'CONDITIONAL'}
    assert 'finance' not in run['financial_model']
    assert 'curve_type' not in run['financial_model']['sales']
    assert run['summary']['total_developer_equity_commitment'] == '12000000'
    assert run['summary']['initial_equity_contribution'] == '12000000'
    assert {item['method'] for item in run['negotiation_results']} == {
        'GROSS_SALES', 'NET_SALES', 'PROFIT_SHARE', 'UPFRONT', 'HYBRID', 'MINIMUM_GUARANTEE',
    }

    with session_scope() as db:
        stored = db.get(CalculationRun, run['id'])
        assert stored and stored.input_hash == run['input_hash'] and stored.result_hash == run['result_hash']
        assert db.scalar(select(CalculationRunResult).where(CalculationRunResult.calculation_run_id == run['id']))
        assert len(list(db.scalars(select(NegotiationResult).where(NegotiationResult.calculation_run_id == run['id'])))) == 6
        assert len(list(db.scalars(select(MonthlyCashflowSnapshot).where(MonthlyCashflowSnapshot.calculation_run_id == run['id'])))) == len(run['monthly_cashflow'])

    pdf = client.get(f"/api/projects/{project_id}/financial/runs/{run['id']}/report.pdf")
    assert pdf.status_code == 200, pdf.text
    assert pdf.content.startswith(b'%PDF') and len(pdf.content) > 5000

    excel = client.get(f"/api/projects/{project_id}/financial/runs/{run['id']}/report.xlsx")
    assert excel.status_code == 200, excel.text
    assert excel.content.startswith(b'PK')
    workbook = load_workbook(BytesIO(excel.content), data_only=False)
    assert workbook.sheetnames == [
        'Executive Summary', 'Negotiation Range', 'Annual Cash Flow',
        'Monthly Cash Flow', 'Inputs and Provenance',
    ]
    assert workbook['Executive Summary']['A1'].value == 'LandValue360 - Standalone Financial Portal'


def test_financial_policy_versioning_is_immutable(client):
    register(client, email='policy-owner@example.com', org='Policy Owner')
    with session_scope() as db:
        create_staff_user(
            db, email='policy-admin@example.com', password='StrongPass123!',
            full_name='Policy Admin', role_code='PLATFORM_ADMIN',
        )
    owner_csrf = client.get('/api/auth/me').json()['csrf_token']
    client.post('/api/auth/logout', headers={'X-CSRF-Token': owner_csrf})
    csrf = login(client, 'policy-admin@example.com')
    current = client.get('/api/admin/financial-policy')
    assert current.status_code == 200, current.text
    first = current.json()['current']
    controls = first['controls']
    controls['target_developer_irr'] = '0.24'
    controls['proposal_selection_method'] = 'MAXIMUM_LANDOWNER_VALUE'
    controls['advanced_defaults']['finance_enabled'] = False
    controls['advanced_defaults']['annual_interest_rate'] = '0.09'
    controls['advanced_defaults']['sales_curve_type'] = 'S_CURVE'
    controls['advanced_defaults']['construction_curve_type'] = 'BELL'
    response = client.post('/api/admin/financial-policy/versions', json={
        'controls': controls,
        'change_reason': 'Raise target return and maximize landowner value within constraints',
    }, headers={'X-CSRF-Token': csrf})
    assert response.status_code == 201, response.text
    second = response.json()
    assert second['version_number'] == first['version_number'] + 1
    assert second['snapshot_hash'] != first['snapshot_hash']
    assert second['controls']['target_developer_irr'] == '0.24'
    assert second['controls']['proposal_selection_method'] == 'MAXIMUM_LANDOWNER_VALUE'
    assert second['controls']['advanced_defaults']['finance_enabled'] is False
    assert second['controls']['advanced_defaults']['annual_interest_rate'] == '0.09'
    assert second['controls']['advanced_defaults']['sales_curve_type'] == 'S_CURVE'
    assert second['controls']['advanced_defaults']['construction_curve_type'] == 'BELL'
    with session_scope() as db:
        rows = list(db.scalars(select(FinancialPolicyVersion).order_by(FinancialPolicyVersion.version_number)).all())
        assert len(rows) == 2
        assert all(row.immutable for row in rows)
        assert rows[0].snapshot_hash == first['snapshot_hash']


def test_legacy_project_defaults_are_frozen_inside_input_hash(client):
    project_id, csrf = setup_project(client)
    with session_scope() as db:
        project = db.get(Project, project_id)
        version = db.get(ProjectVersion, project.current_version_id)
        snapshot = dict(version.input_snapshot or {})
        snapshot.pop('financial_model', None)
        version.input_snapshot = snapshot
        version.snapshot_hash = sha256_json(snapshot)
        source_hash = version.snapshot_hash

    first_response = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={},
        headers={'X-CSRF-Token': csrf},
    )
    assert first_response.status_code == 201, first_response.text
    first = first_response.json()
    second_response = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={},
        headers={'X-CSRF-Token': csrf},
    )
    assert second_response.status_code == 201, second_response.text
    second = second_response.json()
    assert first['input_hash'] == second['input_hash']
    assert first['result_hash'] == second['result_hash']

    with session_scope() as db:
        stored = db.get(CalculationRun, first['id'])
        result = db.scalar(select(CalculationRunResult).where(CalculationRunResult.calculation_run_id == first['id']))
        frozen = stored.input_snapshot['project_snapshot']['financial_model']
        assert stored.input_snapshot['source_project_snapshot_hash'] == source_hash
        assert frozen['valuation_date']
        assert frozen['finance']['allow_negative_cash'] is False
        assert stored.input_hash == sha256_json(stored.input_snapshot)
        assert result.full_result['financial_model'] == frozen


def test_calculation_rejects_non_executable_engine_record(client):
    project_id, csrf = setup_project(client)
    with session_scope() as db:
        engine = db.scalar(select(EngineVersion).where(EngineVersion.active.is_(True)))
        engine.source_hash = '0' * 64

    response = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={},
        headers={'X-CSRF-Token': csrf},
    )
    assert response.status_code == 422, response.text
    assert 'not executable' in response.json()['detail']


def test_run_provenance_remains_frozen_after_draft_version_changes(client):
    project_id, csrf = setup_project(client)
    state = client.get(f'/api/projects/{project_id}/financial').json()
    model = state['financial_model']
    model['funding']['opening_cash'] = '12000000'
    model['funding']['total_developer_equity'] = '12000000'
    saved = client.put(
        f'/api/projects/{project_id}/financial', json=model, headers={'X-CSRF-Token': csrf},
    )
    assert saved.status_code == 200, saved.text
    source_hash = saved.json()['snapshot_hash']

    created = client.post(
        f'/api/projects/{project_id}/financial/runs', json={}, headers={'X-CSRF-Token': csrf},
    )
    assert created.status_code == 201, created.text
    run = created.json()
    frozen_effective_hash = run['effective_project_input_hash']
    assert run['source_project_snapshot_hash'] == source_hash
    assert run['current_project_snapshot_hash'] == source_hash

    model['sales']['commercial_discount_rate'] = '0.05'
    changed = client.put(
        f'/api/projects/{project_id}/financial', json=model, headers={'X-CSRF-Token': csrf},
    )
    assert changed.status_code == 200, changed.text
    changed_hash = changed.json()['snapshot_hash']
    assert changed_hash != source_hash

    retrieved = client.get(f"/api/projects/{project_id}/financial/runs/{run['id']}")
    assert retrieved.status_code == 200, retrieved.text
    payload = retrieved.json()
    assert payload['project_snapshot_hash'] == source_hash
    assert payload['source_project_snapshot_hash'] == source_hash
    assert payload['effective_project_input_hash'] == frozen_effective_hash
    assert payload['current_project_snapshot_hash'] == changed_hash
    assert payload['project_input_snapshot_frozen'] is True


def test_policy_and_financial_inputs_use_canonical_decimal_rates():
    policy = default_financial_policy_snapshot()
    controls = policy['portal_policy']
    controls['minimum_profit_on_cost'] = '1.5'
    controls['target_developer_profit_on_cost'] = '2'
    normalized = apply_policy_controls(policy, controls)
    assert normalized['portal_policy']['minimum_profit_on_cost'] == '1.5'
    assert normalized['portal_policy']['target_developer_profit_on_cost'] == '2'

    invalid = dict(controls)
    invalid['target_developer_irr'] = '0.10'
    invalid['minimum_developer_equity_irr'] = '0.20'
    try:
        apply_policy_controls(policy, invalid)
    except ValueError as exc:
        assert 'cannot be below' in str(exc)
    else:
        raise AssertionError('Conflicting IRR policy thresholds must be rejected')

    raw = {'advanced_overrides_enabled': True, 'sales': {'collection_rules': [{'lag_months': 0, 'weight': '-0.2'}]}}
    try:
        normalize_financial_model(raw)
    except ValueError as exc:
        assert 'cannot be negative' in str(exc)
    else:
        raise AssertionError('Negative collection weights must be rejected')


def test_v21_equity_commitment_includes_initial_contribution():
    model = normalize_financial_model({
        'funding': {'opening_cash': '3000000', 'total_developer_equity': '10000000'},
    })
    assert model['funding']['opening_cash'] == '3000000'
    assert model['funding']['total_developer_equity'] == '10000000'
    assert model['funding']['committed_additional_equity'] == '7000000'


def test_v21_policy_managed_advanced_defaults_are_unlevered_and_logical():
    controls = default_financial_policy_snapshot()['portal_policy']
    model = normalize_financial_model({}, controls=controls)
    assert model['advanced_overrides_enabled'] is False
    assert model['finance']['enabled'] is False
    assert model['funding']['committed_financing'] == '0'
    assert model['sales']['curve_type'] == 'S_CURVE'
    assert model['delivery']['construction_curve_type'] == 'BELL'
    assert model['delivery']['other_cost_curve_type'] == 'BELL'
    assert [row['weight'] for row in model['sales']['collection_rules']] == ['0.2', '0.3', '0.5']


def test_v21_analyst_override_can_change_advanced_inputs_without_negative_cash():
    controls = default_financial_policy_snapshot()['portal_policy']
    model = normalize_financial_model({
        'advanced_overrides_enabled': True,
        'sales': {
            'curve_type': 'LINEAR',
            'collection_rules': [
                {'lag_months': 0, 'weight': '0.5'},
                {'lag_months': 6, 'weight': '0.5'},
            ],
        },
        'funding': {'opening_cash': '3000000', 'total_developer_equity': '10000000', 'committed_financing': '5000000'},
        'finance': {'enabled': True, 'annual_interest_rate': '0.09', 'allow_negative_cash': True},
    }, controls=controls)
    assert model['advanced_overrides_enabled'] is True
    assert model['sales']['curve_type'] == 'LINEAR'
    assert model['finance']['enabled'] is True
    assert model['funding']['committed_financing'] == '5000000'
    assert model['finance']['annual_interest_rate'] == '0.09'
    assert model['finance']['allow_negative_cash'] is False


def test_analyst_role_exposes_advanced_financial_permission(client):
    register(client, email='role-owner@example.com', org='Role Owner')
    with session_scope() as db:
        create_staff_user(
            db, email='advanced-analyst@example.com', password='StrongPass123!',
            full_name='Advanced Analyst', role_code='ANALYST',
        )
    owner_csrf = client.get('/api/auth/me').json()['csrf_token']
    client.post('/api/auth/logout', headers={'X-CSRF-Token': owner_csrf})
    login(client, 'advanced-analyst@example.com')
    me = client.get('/api/auth/me')
    assert me.status_code == 200
    assert 'financial.advanced_inputs' in me.json()['permissions']


def test_standard_user_can_select_published_policy_but_cannot_override_engine_or_read_full_payload(client):
    project_id, csrf = setup_project(client)
    state = client.get(f'/api/projects/{project_id}/financial').json()
    selected_policy = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={'policy_version_id': state['policy']['id']},
        headers={'X-CSRF-Token': csrf},
    )
    assert selected_policy.status_code == 201, selected_policy.text
    assert selected_policy.json()['financial_policy_version_id'] == state['policy']['id']
    blocked_engine = client.post(
        f'/api/projects/{project_id}/financial/runs',
        json={'engine_version_id': state['engine']['id']},
        headers={'X-CSRF-Token': csrf},
    )
    assert blocked_engine.status_code == 403, blocked_engine.text

    created = client.post(
        f'/api/projects/{project_id}/financial/runs', json={}, headers={'X-CSRF-Token': csrf},
    )
    assert created.status_code == 201, created.text
    run_id = created.json()['id']
    full = client.get(f'/api/projects/{project_id}/financial/runs/{run_id}?include_full=true')
    assert full.status_code == 403, full.text
    audit = client.get(f'/api/projects/{project_id}/financial/runs/{run_id}/audit')
    assert audit.status_code == 200, audit.text
    assert audit.json()['calculation_run_id'] == run_id
    assert audit.json()['financial_audit']['validation_status'] in {'VALIDATED', 'CONDITIONAL'}


def test_standard_user_edit_preserves_existing_advanced_override_server_side(client):
    project_id, csrf = setup_project(client)
    with session_scope() as db:
        project = db.get(Project, project_id)
        version = db.get(ProjectVersion, project.current_version_id)
        snapshot = dict(version.input_snapshot or {})
        model = normalize_financial_model({
            'advanced_overrides_enabled': True,
            'sales': {'curve_type': 'LINEAR'},
            'funding': {'opening_cash': '1000000', 'total_developer_equity': '3000000', 'committed_financing': '5000000'},
            'finance': {'enabled': True, 'annual_interest_rate': '0.11'},
        })
        snapshot['financial_model'] = model
        version.input_snapshot = snapshot
        version.snapshot_hash = sha256_json(snapshot)

    state = client.get(f'/api/projects/{project_id}/financial').json()
    assert 'finance' not in state['financial_model']
    state['financial_model']['funding']['opening_cash'] = '1500000'
    saved = client.put(
        f'/api/projects/{project_id}/financial',
        json=state['financial_model'],
        headers={'X-CSRF-Token': csrf},
    )
    assert saved.status_code == 200, saved.text
    with session_scope() as db:
        project = db.get(Project, project_id)
        version = db.get(ProjectVersion, project.current_version_id)
        stored = version.input_snapshot['financial_model']
        assert stored['advanced_overrides_enabled'] is True
        assert stored['sales']['curve_type'] == 'LINEAR'
        assert stored['finance']['enabled'] is True
        assert stored['finance']['annual_interest_rate'] == '0.11'
        assert stored['funding']['committed_financing'] == '5000000'
        assert stored['funding']['opening_cash'] == '1500000'
