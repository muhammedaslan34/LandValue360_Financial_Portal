#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARTIFACTS = ROOT / "release_artifacts"
PDF = ARTIFACTS / "LandValue360_Financial_Portal_Sample_Report_AR.pdf"
XLSX = ARTIFACTS / "LandValue360_Financial_Portal_Sample_Cashflow.xlsx"
EXPECTED_SHEETS = [
    "Executive Summary",
    "Negotiation Range",
    "Annual Cash Flow",
    "Monthly Cash Flow",
    "Inputs and Provenance",
]
ERROR_VALUES = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}


def main() -> int:
    errors: list[str] = []
    pdf_report: dict[str, object] = {"path": PDF.name}
    if not PDF.is_file() or PDF.read_bytes()[:5] != b"%PDF-":
        errors.append("Financial PDF is missing or invalid")
    else:
        pdf_report["bytes"] = PDF.stat().st_size
        pdf_report["sha256"] = hashlib.sha256(PDF.read_bytes()).hexdigest()
        if shutil.which("pdfinfo"):
            proc = subprocess.run(["pdfinfo", str(PDF)], capture_output=True, text=True)
            pdf_report["pdfinfo_returncode"] = proc.returncode
            metadata: dict[str, str] = {}
            for line in proc.stdout.splitlines():
                if ":" in line:
                    key, value = line.split(":", 1)
                    metadata[key.strip()] = value.strip()
            pdf_report["metadata"] = metadata
            try:
                pages = int(metadata.get("Pages", "0"))
            except ValueError:
                pages = 0
            pdf_report["pages"] = pages
            if proc.returncode != 0 or pages < 1:
                errors.append("Financial PDF metadata validation failed")
        else:
            pdf_report["pdfinfo"] = "unavailable"

    xlsx_report: dict[str, object] = {"path": XLSX.name}
    if not XLSX.is_file():
        errors.append("Financial Excel report is missing")
    else:
        workbook = load_workbook(XLSX, read_only=False, data_only=False)
        xlsx_report["sheets"] = workbook.sheetnames
        if workbook.sheetnames != EXPECTED_SHEETS:
            errors.append("Financial Excel sheet contract changed")
        formula_count = 0
        error_cells: list[str] = []
        sheet_stats: list[dict[str, object]] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                    if value in ERROR_VALUES:
                        error_cells.append(f"{worksheet.title}!{cell.coordinate}")
            sheet_stats.append({
                "name": worksheet.title,
                "rows": worksheet.max_row,
                "columns": worksheet.max_column,
                "freeze_panes": str(worksheet.freeze_panes or ""),
                "gridlines": bool(worksheet.sheet_view.showGridLines),
            })
        if error_cells:
            errors.append(f"Excel contains formula error values: {len(error_cells)}")
        if getattr(workbook, "_external_links", []):
            errors.append("Excel contains external links")
        xlsx_report.update({
            "bytes": XLSX.stat().st_size,
            "sha256": hashlib.sha256(XLSX.read_bytes()).hexdigest(),
            "formula_cells": formula_count,
            "error_cells": error_cells,
            "external_links": len(getattr(workbook, "_external_links", [])),
            "sheet_stats": sheet_stats,
        })

    report = {
        "status": "PASS" if not errors else "FAIL",
        "pdf": pdf_report,
        "xlsx": xlsx_report,
        "errors": errors,
    }
    output = ARTIFACTS / "report-artifacts-validation.json"
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (ARTIFACTS / "spreadsheet-qa.json").write_text(
        json.dumps({"status": report["status"], **xlsx_report, "errors": errors}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
